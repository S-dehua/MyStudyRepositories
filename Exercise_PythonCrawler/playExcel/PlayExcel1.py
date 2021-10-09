import openpyxl
import datetime
wb = openpyxl.Workbook()
ws = wb.active
ws.title
ws['A1'] = 520
ws.append([1,2,3])
ws['A3'] = datetime.datetime.now()
wb.save("demo.xlsx")

